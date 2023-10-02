import re
from openpyxl import Workbook
from openpyxl.styles import Font

def parse_header_file(header_file_path):
    fun_prototypes = []
    
    with open(header_file_path, 'r') as file:
        content = file.read()

        # Use regular expression to find function prototypes
        function_prototypes = re.findall(r'\w+\s+\w+\([^;]*\);', content)

        for idx, prototype in enumerate(function_prototypes):
            # Add unique ID to each prototype
            unique_id = f"IDX{idx}"
            fun_prototypes.append((unique_id, prototype))

    return fun_prototypes

def write_to_excel(prototypes, excel_file_path):
    workbook = Workbook()
    sheet = workbook.active

    # Add headers
    sheet['A1'] = 'Unique ID'
    sheet['B1'] = 'Function Prototype'

    # Apply bold font to headers
    for cell in sheet['1:1']:
        cell.font = Font(bold=True)

    # Populate Excel sheet with function prototypes
    for row_idx, (unique_id, prototype) in enumerate(prototypes, start=2):
        sheet.cell(row=row_idx, column=1, value=unique_id)
        sheet.cell(row=row_idx, column=2, value=prototype)

    # Save Excel file
    workbook.save(excel_file_path)


header_file_path = "/home/muhammed/Documents/python_exp/my_file.h"
excel_file_path = "/home/muhammed/Documents/python_exp/output.xlsx"

function_prototypes = parse_header_file(header_file_path)
write_to_excel(function_prototypes, excel_file_path)
